from openpyxl import Workbook, load_workbook

def crear_excel(nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    wb.save(nombre_archivo)
    print("Archivo creado:", nombre_archivo)

def escribir_datos(nombre_archivo, datos):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    for fila in datos:
        ws.append(fila)
    wb.save(nombre_archivo)
    print("Datos escritos en:", nombre_archivo)

def leer_datos(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    for fila in ws.iter_rows(values_only=True):
        print(fila)
        #samuelpuerto3203084